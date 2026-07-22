from openpyxl import load_workbook
from datetime import datetime
import os

# Path to the Excel file
EXCEL_PATH = os.path.join("data", "timetable.xlsx")


def get_today_schedule():
    wb = load_workbook(EXCEL_PATH, data_only=True)

    calendar = wb["Academic_Calendar"]
    weekly = wb["Weekly_Timetable"]

    today = datetime.now().date()

    status = None
    timetable_day = None

    # Find today's entry
    for row in calendar.iter_rows(min_row=2, values_only=True):
        excel_date = row[0]

        # Convert datetime to date
        if hasattr(excel_date, "date"):
            excel_date = excel_date.date()

        if excel_date == today:
            status = row[2]
            timetable_day = row[3]
            break

    # If today's date wasn't found
    if timetable_day is None:
        return {
            "date": str(today),
            "status": "No Entry",
            "timetable": None,
            "subjects": []
        }

    # Holiday
    if status == "Holiday":
        return {
            "date": str(today),
            "status": "Holiday",
            "timetable": timetable_day,
            "subjects": []
        }

    # Get subjects
    subjects = []

    for row in weekly.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().lower() == str(timetable_day).strip().lower():
            subjects = [x for x in row[1:] if x]
            break

    return {
        "date": str(today),
        "status": "Working",
        "timetable": timetable_day,
        "subjects": subjects
    }